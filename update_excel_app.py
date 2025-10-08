"""
☕ Café Cultura — Excel Updater
A Streamlit app to update Excel files based on training completion status.
© 2025 Rodrigo Bermudez — Café Cultura
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Page configuration
st.set_page_config(
    page_title="Café Cultura — Excel Updater",
    page_icon="☕",
    layout="wide"
)

# Custom CSS for a clean, minimalist theme
st.markdown("""
    <style>
    .main-header {
        text-align: center;
        padding: 1rem;
        background: linear-gradient(90deg, #6B4423 0%, #8B6F47 100%);
        color: white;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .footer {
        text-align: center;
        padding: 2rem;
        color: #666;
        font-size: 0.9rem;
        margin-top: 3rem;
        border-top: 1px solid #ddd;
    }
    </style>
""", unsafe_allow_html=True)

# Header
st.markdown('<div class="main-header"><h1>☕ Café Cultura — Excel Updater</h1></div>', unsafe_allow_html=True)

# Introduction
st.markdown("""
Welcome to the **Café Cultura Excel Updater**! This tool helps you automatically update training completion
status in your main Excel file based on export data.
""")

# Sidebar instructions
with st.sidebar:
    st.header("📋 Instructions")
    st.markdown("""
    1. **Upload** your main Excel file
    2. **Upload** your export file (with Status column)
    3. **Select** the training column to update
    4. **Define** the training columns range
    5. **Process** and download the updated file

    ---

    ### 💡 Tips
    - Names are matched using First + Last Name
    - Only "Completed" status updates the training
    - Rows with all trainings complete are highlighted in yellow
    """)

# Main content
col1, col2 = st.columns(2)

with col1:
    st.subheader("📁 Upload Main File")
    main_file = st.file_uploader(
        "Choose your main Excel file",
        type=['xlsx', 'xls'],
        key='main_file',
        help="This is the file that will be updated"
    )

with col2:
    st.subheader("📁 Upload Export File")
    export_file = st.file_uploader(
        "Choose your export Excel file",
        type=['xlsx', 'xls'],
        key='export_file',
        help="This file should contain a 'Status' column"
    )

# Function to read Excel files with caching
@st.cache_data
def load_excel_file(file_bytes, file_name):
    """Load Excel file from uploaded bytes"""
    return pd.read_excel(BytesIO(file_bytes))

# Process files if both are uploaded
if main_file and export_file:
    try:
        # Read the Excel files
        with st.spinner("📖 Reading Excel files..."):
            main_df = load_excel_file(main_file.getvalue(), main_file.name)
            export_df = load_excel_file(export_file.getvalue(), export_file.name)

        st.success("✅ Files loaded successfully!")

        # Display file previews
        with st.expander("👀 Preview Main File"):
            st.dataframe(main_df.head(10))

        with st.expander("👀 Preview Export File"):
            st.dataframe(export_df.head(10))

        # Check for required columns
        st.subheader("⚙️ Configuration")

        # Identify name columns in main file
        main_cols = main_df.columns.tolist()

        col_config1, col_config2 = st.columns(2)

        with col_config1:
            first_name_col_main = st.selectbox(
                "📝 First Name Column (Main File)",
                main_cols,
                index=next((i for i, col in enumerate(main_cols) if 'first' in col.lower()), 0)
            )

            last_name_col_main = st.selectbox(
                "📝 Last Name Column (Main File)",
                main_cols,
                index=next((i for i, col in enumerate(main_cols) if 'last' in col.lower()), 1)
            )

        with col_config2:
            # Identify name columns in export file
            export_cols = export_df.columns.tolist()

            first_name_col_export = st.selectbox(
                "📝 First Name Column (Export File)",
                export_cols,
                index=next((i for i, col in enumerate(export_cols) if 'first' in col.lower()), 0)
            )

            last_name_col_export = st.selectbox(
                "📝 Last Name Column (Export File)",
                export_cols,
                index=next((i for i, col in enumerate(export_cols) if 'last' in col.lower()), 1)
            )

        # Check for Status column in export file
        status_col = st.selectbox(
            "✅ Status Column (Export File)",
            export_cols,
            index=next((i for i, col in enumerate(export_cols) if 'status' in col.lower()), 0)
        )

        # Select training column to update
        training_column = st.selectbox(
            "🎯 Select Training Column to Update",
            main_cols,
            help="Choose which training column should be updated with 'Completed' status"
        )

        # Define training columns range for highlighting
        st.markdown("### 🎨 Highlighting Configuration")
        st.info("Define the range of training columns. Rows where ALL these trainings are 'Completed' will be highlighted in yellow.")

        range_col1, range_col2 = st.columns(2)

        with range_col1:
            start_col = st.selectbox(
                "📍 Start Column",
                main_cols,
                help="First training column in the range"
            )

        with range_col2:
            end_col = st.selectbox(
                "📍 End Column",
                main_cols,
                index=len(main_cols)-1,
                help="Last training column in the range"
            )

        # Process button
        if st.button("🚀 Process Files", type="primary", use_container_width=True):
            with st.spinner("⏳ Processing your files..."):
                # Create full name columns for matching (case-insensitive)
                main_df['_full_name'] = (
                    main_df[first_name_col_main].astype(str).str.strip().str.lower() + ' ' +
                    main_df[last_name_col_main].astype(str).str.strip().str.lower()
                )

                export_df['_full_name'] = (
                    export_df[first_name_col_export].astype(str).str.strip().str.lower() + ' ' +
                    export_df[last_name_col_export].astype(str).str.strip().str.lower()
                )

                # Filter export data for completed status
                completed_df = export_df[export_df[status_col].astype(str).str.strip().str.lower() == 'completed']

                # Track updates
                updates_count = 0

                # Update the training column for matched names
                for idx, row in main_df.iterrows():
                    full_name = row['_full_name']
                    if full_name in completed_df['_full_name'].values:
                        main_df.at[idx, training_column] = 'Completed'
                        updates_count += 1

                # Remove helper column
                main_df = main_df.drop('_full_name', axis=1)

                st.success(f"✅ Updated {updates_count} records in the '{training_column}' column!")

                # Save to Excel with highlighting
                output = BytesIO()

                # Get column indices for the range
                start_idx = main_df.columns.get_loc(start_col)
                end_idx = main_df.columns.get_loc(end_col)
                training_cols = main_df.columns[start_idx:end_idx+1].tolist()

                # Write to Excel first
                main_df.to_excel(output, index=False, engine='openpyxl')
                output.seek(0)

                # Apply highlighting using openpyxl
                wb = load_workbook(output)
                ws = wb.active

                # Yellow fill for completed rows
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                # Track highlighted rows
                highlighted_count = 0

                # Check each row (starting from row 2, as row 1 is header)
                for row_idx in range(2, len(main_df) + 2):
                    # Check if all training columns in range are "Completed"
                    all_completed = True
                    for col_name in training_cols:
                        col_idx = main_df.columns.get_loc(col_name) + 1  # +1 for Excel 1-based indexing
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if str(cell_value).strip().lower() != 'completed':
                            all_completed = False
                            break

                    # Highlight entire row if all trainings are completed
                    if all_completed:
                        for col_idx in range(1, len(main_df.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
                        highlighted_count += 1

                # Save the workbook
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                st.success(f"🎨 Highlighted {highlighted_count} rows where all trainings are completed!")

                # Display preview
                st.subheader("📊 Updated Data Preview")
                st.dataframe(main_df.head(20))

                # Download button
                st.download_button(
                    label="⬇️ Download Updated Excel File",
                    data=output,
                    file_name="updated_training_file.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )

                # Statistics
                st.subheader("📈 Statistics")
                stats_col1, stats_col2, stats_col3 = st.columns(3)

                with stats_col1:
                    st.metric("Total Records", len(main_df))

                with stats_col2:
                    st.metric("Updated Records", updates_count)

                with stats_col3:
                    st.metric("Highlighted Rows", highlighted_count)

    except Exception as e:
        st.error(f"❌ Error processing files: {str(e)}")
        st.info("💡 Please check that your files are properly formatted and contain the expected columns.")

else:
    st.info("👆 Please upload both Excel files to get started!")

# Footer
st.markdown("""
<div class="footer">
    <strong>© 2025 Rodrigo Bermudez — Café Cultura</strong>
</div>
""", unsafe_allow_html=True)
